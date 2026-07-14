from datetime import date
from decimal import Decimal
import unittest

from openpyxl import Workbook

from tools.excel_import.moneyflow_excel_import import (
    NormalCandidate,
    day_from_header,
    ensure_execute_guard,
    normal_key,
    parse_amount,
    parse_date,
    parse_normal_transactions,
    select_workspace,
)


class MoneyFlowExcelImportTests(unittest.TestCase):
    def test_date_parsing(self):
        self.assertEqual(parse_date("01/06/2026"), date(2026, 6, 1))
        self.assertEqual(parse_date("2026-07-31"), date(2026, 7, 31))

    def test_amount_parsing(self):
        self.assertEqual(parse_amount("1,234,500"), Decimal("1234500.00"))
        self.assertIsNone(parse_amount(0))

    def test_normal_key_normalization(self):
        a = NormalCandidate("T6", 57, date(2026, 6, 1), "EXPENSE", Decimal("1000.00"), "Công nợ", "", "", "T6!A1")
        b = NormalCandidate("T6", 57, date(2026, 6, 1), "EXPENSE", Decimal("1000.00"), "Cong no", "", "", "T6!A1")
        self.assertEqual(normal_key(a), normal_key(b))

    def test_workspace_selection_by_exact_name(self):
        class Cursor:
            def execute(self, *_):
                pass

            def fetchall(self):
                return [("workspace-id", "Tài chính cá nhân")]

        ws = select_workspace(Cursor(), "Tài chính cá nhân")
        self.assertEqual(ws.id, "workspace-id")

    def test_fixed_expense_day_headers(self):
        self.assertEqual(day_from_header("Đầu tháng", 2026, 6), date(2026, 6, 1))
        self.assertEqual(day_from_header("Ngày 16", 2026, 7), date(2026, 7, 16))
        self.assertEqual(day_from_header("Cuối tháng", 2026, 7), date(2026, 7, 31))

    def test_parse_normal_transactions_daily_and_fixed(self):
        wb = Workbook()
        wb.active.title = "T6"
        wb.create_sheet("T7")
        for ws in [wb["T6"], wb["T7"]]:
            ws["A55"] = "Thu Nhập Của Anh"
            ws["B55"] = "Thu Nhập Của Em"
            ws["E56"] = "Cafe"
            ws["N45"] = "Đầu tháng"
            ws["N46"] = "Tiền trọ:"
            ws["N47"] = 1000
        wb["T6"]["C57"] = date(2026, 6, 1)
        wb["T6"]["A57"] = 2000
        wb["T6"]["E57"] = 3000
        wb["T7"]["C57"] = date(2026, 7, 1)
        wb["T7"]["B57"] = 4000
        rows = parse_normal_transactions(wb)
        self.assertEqual(len(rows), 5)
        income = sum((r.amount for r in rows if r.tx_type == "INCOME"), Decimal("0.00"))
        self.assertEqual(income, Decimal("6000.00"))

    def test_execute_guard_requires_all_confirmation_flags(self):
        normal = [NormalCandidate("T6", 57, date(2026, 6, 1), "INCOME", Decimal("10.00"), "", "", "", "T6!A1")]
        args = type(
            "Args",
            (),
            {
                "execute": True,
                "i_understand_this_mutates_db": True,
                "confirm_workspace_name": "wrong",
                "workspace_name": "right",
                "confirm_normal_count": 1,
                "confirm_normal_expense_total": Decimal("0.00"),
                "confirm_normal_income_total": Decimal("10.00"),
            },
        )()
        with self.assertRaises(SystemExit):
            ensure_execute_guard(args, normal)

    def test_debt_candidates_not_executable_by_default(self):
        args = type("Args", (), {"execute": False})()
        self.assertIsNone(ensure_execute_guard(args, []))


if __name__ == "__main__":
    unittest.main()
