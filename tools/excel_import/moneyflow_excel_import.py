#!/usr/bin/env python3
"""Guarded dry-run June/July 2026 Excel truth reporter for MoneyFlow."""

from __future__ import annotations

import argparse
import csv
import hashlib
import os
import re
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Iterable
from urllib.parse import urlsplit, urlunsplit

from openpyxl import load_workbook

START = date(2026, 6, 1)
END = date(2026, 8, 1)
MONTH_SHEETS = ("T6", "T7")
REPORT_DIR = Path("reports")

NORMAL_CSV = REPORT_DIR / "legacy-june-july-normal-truth.csv"
NORMAL_MD = REPORT_DIR / "legacy-june-july-normal-truth.md"
DEBT_CSV = REPORT_DIR / "legacy-june-july-debt-truth.csv"
PAYMENT_CSV = REPORT_DIR / "legacy-june-july-debt-payments-truth.csv"
PERSON_MD = REPORT_DIR / "legacy-june-july-debt-person-summary.md"
DB_NORMAL_MD = REPORT_DIR / "legacy-june-july-normal-db-diff.md"
DB_DEBT_MD = REPORT_DIR / "legacy-june-july-debt-db-diff.md"
PLAN_MD = REPORT_DIR / "legacy-june-july-import-plan.md"


@dataclass(frozen=True)
class NormalCandidate:
    sheet: str
    row: int
    tx_date: date
    tx_type: str
    amount: Decimal
    category: str
    wallet: str
    description: str
    raw_cells: str
    confidence: str = "HIGH"
    skip_reason: str = ""


@dataclass(frozen=True)
class DebtCandidate:
    sheet: str
    row: int
    person_name: str
    debt_type: str
    opened_date: date
    due_date: date | None
    pay_date: date | None
    principal_amount: Decimal
    paid_amount: Decimal
    remaining_amount: Decimal
    status: str
    note: str
    raw_cells: str
    confidence: str
    ambiguity_reason: str


@dataclass(frozen=True)
class PaymentCandidate:
    sheet: str
    row: int
    person_name: str
    payment_date: date
    amount: Decimal
    linked_debt_clue: str
    note: str
    raw_cells: str
    confidence: str
    ambiguity_reason: str


@dataclass(frozen=True)
class Workspace:
    id: str
    name: str


def parse_date(value) -> date:
    if value in (None, ""):
        raise ValueError("missing date")
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        text = value.strip()
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(text, fmt).date()
            except ValueError:
                pass
    raise ValueError(f"invalid date: {value!r}")


def parse_amount(value) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        amount = value
    elif isinstance(value, (int, float)):
        amount = Decimal(str(value))
    else:
        text = str(value).strip().replace(" ", "")
        text = re.sub(r"[^0-9,\.\-]", "", text)
        if text.count(",") == 1 and text.count(".") >= 1 and text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "").replace(".", "") if re.fullmatch(r"-?[0-9.]+", text) else text.replace(",", "")
        if not text:
            return None
        try:
            amount = Decimal(text)
        except (InvalidOperation, ValueError) as exc:
            raise ValueError(f"invalid amount: {value!r}") from exc
    try:
        amount = amount.quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"invalid amount: {value!r}") from exc
    return amount if amount != 0 else None


def normalize_text(value: str | None) -> str:
    text = (value or "").strip().lower().replace("đ", "d")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text)


def in_scope(d: date | None) -> bool:
    return d is not None and START <= d < END


def sum_amount(rows: Iterable[NormalCandidate], tx_type: str) -> Decimal:
    return sum((r.amount for r in rows if r.tx_type == tx_type), Decimal("0.00"))


def month_total(rows: Iterable[NormalCandidate], month: int, tx_type: str) -> Decimal:
    return sum((r.amount for r in rows if r.tx_date.month == month and r.tx_type == tx_type), Decimal("0.00"))


def raw(ws, row: int, cols: Iterable[int]) -> str:
    parts = []
    for col in cols:
        cell = ws.cell(row, col)
        if cell.value not in (None, ""):
            parts.append(f"{cell.coordinate}={cell.value}")
    return "; ".join(parts)


def day_from_header(value, year: int, month: int) -> date | None:
    text = str(value or "").strip()
    if normalize_text(text) == "dau thang":
        return date(year, month, 1)
    if normalize_text(text) == "cuoi thang":
        return date(year, month, 30 if month == 6 else 31)
    m = re.search(r"(\d+)", text)
    return date(year, month, int(m.group(1))) if m else None


def workbook_path(value: Path) -> Path:
    if value.exists():
        return value
    matches = [p for p in value.parent.glob("*.xlsx") if p.name == value.name]
    if matches:
        return matches[0]
    raise FileNotFoundError(value)


def inspect_workbook(excel: Path):
    wb = load_workbook(excel, data_only=True)
    rows = []
    for ws in wb.worksheets:
        nonempty = [
            r
            for r in range(1, ws.max_row + 1)
            if any(ws.cell(r, c).value not in (None, "") for c in range(1, ws.max_column + 1))
        ]
        rows.append((ws.title, ws.calculate_dimension(), min(nonempty or [0]), max(nonempty or [0])))
    return wb, rows


def parse_normal_transactions(wb) -> list[NormalCandidate]:
    rows: list[NormalCandidate] = []
    for sheet in MONTH_SHEETS:
        ws = wb[sheet]
        month = 6 if sheet == "T6" else 7
        income_headers = {1: str(ws["A55"].value).strip(), 2: str(ws["B55"].value).strip()}
        expense_headers = {
            col: str(ws.cell(56, col).value).strip()
            for col in range(5, ws.max_column + 1)
            if ws.cell(56, col).value not in (None, "")
        }
        for row in range(57, ws.max_row + 1):
            try:
                tx_date = parse_date(ws.cell(row, 3).value)
            except ValueError:
                continue
            if not in_scope(tx_date):
                continue
            for col, label in income_headers.items():
                amount = parse_amount(ws.cell(row, col).value)
                if amount:
                    rows.append(NormalCandidate(sheet, row, tx_date, "INCOME", amount, label, "", label, raw(ws, row, (col, 3))))
            for col, label in expense_headers.items():
                amount = parse_amount(ws.cell(row, col).value)
                if amount:
                    rows.append(NormalCandidate(sheet, row, tx_date, "EXPENSE", amount, label, "", label, raw(ws, row, (3, col))))
        for col in range(13, 22):
            tx_date = day_from_header(ws.cell(45, col).value, 2026, month)
            amount = parse_amount(ws.cell(47, col).value)
            if tx_date and amount and in_scope(tx_date):
                label = str(ws.cell(46, col).value).strip()
                rows.append(NormalCandidate(sheet, 47, tx_date, "EXPENSE", amount, label, "", label, raw(ws, 47, (col,))))
    return sorted(rows, key=lambda r: (r.tx_date, r.sheet, r.row, r.category))


def debt_sheet(wb):
    return next(ws for ws in wb.worksheets if ws.title.startswith("GHI"))


def status_paid(status: str) -> bool:
    return normalize_text(status) == "da tra"


def parse_debt_truth(wb) -> tuple[list[DebtCandidate], list[PaymentCandidate]]:
    ws = debt_sheet(wb)
    debts: list[DebtCandidate] = []
    payments: list[PaymentCandidate] = []
    for row in range(4, ws.max_row + 1):
        for start_col, debt_type in ((1, "RECEIVABLE"), (8, "PAYABLE")):
            person = str(ws.cell(row, start_col + 1).value or "").strip()
            amount = parse_amount(ws.cell(row, start_col + 3).value)
            if not person or not amount:
                continue
            try:
                opened = parse_date(ws.cell(row, start_col).value)
            except ValueError:
                continue
            due = paid_on = None
            try:
                due = parse_date(ws.cell(row, start_col + 2).value)
            except ValueError:
                pass
            try:
                paid_on = parse_date(ws.cell(row, start_col + 4).value)
            except ValueError:
                pass
            status = str(ws.cell(row, start_col + 5).value or "").strip()
            paid = amount if status_paid(status) else Decimal("0.00")
            remaining = Decimal("0.00") if status_paid(status) else amount
            reasons = []
            if status_paid(status) and not paid_on:
                reasons.append("status paid but payment date is blank")
            if not in_scope(opened) and in_scope(paid_on):
                reasons.append("opened outside June/July; included because payment date is in scope")
            if in_scope(opened) or in_scope(paid_on):
                debts.append(
                    DebtCandidate(
                        ws.title, row, person, debt_type, opened, due, paid_on, amount, paid, remaining, status, "",
                        raw(ws, row, range(start_col, start_col + 6)),
                        "MEDIUM" if reasons else "HIGH",
                        "; ".join(reasons),
                    )
                )
            if paid_on and in_scope(paid_on):
                payments.append(
                    PaymentCandidate(
                        ws.title, row, person, paid_on, amount,
                        f"{debt_type} opened {opened.isoformat()} amount {amount}", "",
                        raw(ws, row, range(start_col, start_col + 6)),
                        "HIGH",
                        "" if in_scope(opened) else "debt opened outside June/July",
                    )
                )
    return debts, payments


def normal_key(c: NormalCandidate) -> tuple[str, str, str, str]:
    return (c.tx_date.isoformat(), str(c.amount), c.tx_type, normalize_text(c.category))


def ensure_execute_guard(args: argparse.Namespace, normal: list[NormalCandidate]) -> None:
    if not args.execute:
        return
    required = [
        args.i_understand_this_mutates_db,
        args.confirm_workspace_name == args.workspace_name,
        args.confirm_normal_count == len(normal),
        args.confirm_normal_expense_total == sum_amount(normal, "EXPENSE"),
        args.confirm_normal_income_total == sum_amount(normal, "INCOME"),
    ]
    if not all(required):
        raise SystemExit("execute refused: all confirmation flags must match the dry-run plan")


def jdbc_to_pg_url(url: str) -> str:
    return "postgresql://" + url[len("jdbc:postgresql://") :] if url.startswith("jdbc:postgresql://") else url


def masked_url(url: str) -> str:
    parsed = urlsplit(jdbc_to_pg_url(url))
    host = parsed.hostname or ""
    port = f":{parsed.port}" if parsed.port else ""
    return urlunsplit((parsed.scheme, f"{host}{port}", parsed.path, "", ""))


def select_workspace(cur, workspace_name: str) -> Workspace:
    cur.execute("SELECT id, name FROM workspaces WHERE name = %s AND deleted_at IS NULL", (workspace_name,))
    rows = cur.fetchall()
    if len(rows) != 1:
        raise SystemExit(f"workspace selection failed: expected 1 exact name match, got {len(rows)}")
    return Workspace(str(rows[0][0]), str(rows[0][1]))


def skipped_db(normal, debts, payments, reason: str):
    people = sorted({d.person_name for d in debts} | {p.person_name for p in payments}, key=normalize_text)
    return {"done": False, "notes": [reason], "exact_normal": [], "missing_normal": normal, "likely_normal": [], "db_people": [], "excel_people": people, "excel_people_missing": [], "db_people_missing": [], "suspicious": []}


def db_compare(workspace_name: str, normal: list[NormalCandidate], debts: list[DebtCandidate], payments: list[PaymentCandidate]):
    url = os.getenv("MONEYFLOW_DB_URL")
    user = os.getenv("MONEYFLOW_DB_USERNAME")
    password = os.getenv("MONEYFLOW_DB_PASSWORD")
    if not (url and user and password):
        return skipped_db(normal, debts, payments, "DB comparison skipped: MONEYFLOW_DB_URL/USERNAME/PASSWORD not all set.")
    try:
        import psycopg  # type: ignore
    except ImportError:
        try:
            import psycopg2 as psycopg  # type: ignore
        except ImportError:
            return skipped_db(normal, debts, payments, f"DB comparison skipped: psycopg/psycopg2 is not installed. DB URL: {masked_url(url)}")
    with psycopg.connect(jdbc_to_pg_url(url), user=user, password=password) as conn:
        try:
            conn.readonly = True
        except Exception:
            pass
        with conn.cursor() as cur:
            cur.execute("BEGIN READ ONLY")
            ws = select_workspace(cur, workspace_name)
            cur.execute(
                """
                SELECT t.transaction_date, t.amount, t.transaction_type, COALESCE(c.name, '')
                FROM transactions t
                LEFT JOIN categories c ON c.id = t.category_id
                WHERE t.workspace_id = %s AND t.transaction_date >= DATE '2026-06-01'
                  AND t.transaction_date < DATE '2026-08-01' AND t.deleted_at IS NULL
                """,
                (ws.id,),
            )
            existing_normal = {
                (parse_date(r[0]).isoformat(), str(Decimal(str(r[1])).quantize(Decimal("0.01"))), str(r[2]), normalize_text(str(r[3] or "")))
                for r in cur.fetchall()
            }
            cur.execute(
                """
                SELECT p.display_name, d.note
                FROM debts d
                JOIN workspace_people p ON p.id = d.counterparty_person_id
                WHERE d.workspace_id = %s
                """,
                (ws.id,),
            )
            debt_rows = cur.fetchall()
            cur.execute("ROLLBACK")
    exact = [c for c in normal if normal_key(c) in existing_normal]
    missing = [c for c in normal if normal_key(c) not in existing_normal]
    excel_people = sorted({d.person_name for d in debts} | {p.person_name for p in payments}, key=normalize_text)
    db_people = sorted({str(r[0]).strip() for r in debt_rows if str(r[0]).strip()}, key=normalize_text)
    db_norm = {normalize_text(p) for p in db_people}
    excel_norm = {normalize_text(p) for p in excel_people}
    return {
        "done": True,
        "notes": [f"DB URL: {masked_url(url)}", f"Workspace: {workspace_name}"],
        "exact_normal": exact,
        "missing_normal": missing,
        "likely_normal": [c for c in missing if any(k[:3] == normal_key(c)[:3] for k in existing_normal)],
        "db_people": db_people,
        "excel_people": excel_people,
        "excel_people_missing": [p for p in excel_people if normalize_text(p) not in db_norm],
        "db_people_missing": [p for p in db_people if normalize_text(p) not in excel_norm],
        "suspicious": [r for r in debt_rows if "Debt import not part of T1" in str(r[1])],
    }


def write_normal_csv(rows: list[NormalCandidate]) -> None:
    with NORMAL_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["sheet", "row", "date", "type", "amount", "categoryText", "walletText", "descriptionNote", "rawCellsUsed", "confidence", "skipReason"])
        for r in rows:
            writer.writerow([r.sheet, r.row, r.tx_date.isoformat(), r.tx_type.lower(), r.amount, r.category, r.wallet, r.description, r.raw_cells, r.confidence, r.skip_reason])


def write_debt_csv(rows: list[DebtCandidate]) -> None:
    with DEBT_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["sheet", "row", "personName", "type", "openedDate", "dueDate", "payDate", "principalAmount", "paidAmount", "remainingAmount", "status", "note", "rawCellsUsed", "confidence", "ambiguityReason"])
        for r in rows:
            writer.writerow([r.sheet, r.row, r.person_name, r.debt_type, r.opened_date.isoformat(), r.due_date.isoformat() if r.due_date else "", r.pay_date.isoformat() if r.pay_date else "", r.principal_amount, r.paid_amount, r.remaining_amount, r.status, r.note, r.raw_cells, r.confidence, r.ambiguity_reason])


def write_payment_csv(rows: list[PaymentCandidate]) -> None:
    with PAYMENT_CSV.open("w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["sheet", "row", "personName", "paymentDate", "amount", "linkedDebtClue", "note", "rawCellsUsed", "confidence", "ambiguityReason"])
        for r in rows:
            writer.writerow([r.sheet, r.row, r.person_name, r.payment_date.isoformat(), r.amount, r.linked_debt_clue, r.note, r.raw_cells, r.confidence, r.ambiguity_reason])


def write_markdown(excel: Path, sheet_info, normal: list[NormalCandidate], debts: list[DebtCandidate], payments: list[PaymentCandidate], db) -> None:
    sha = hashlib.sha256(excel.read_bytes()).hexdigest()
    by_cat = Counter(r.category for r in normal)
    by_day = Counter(r.tx_date.isoformat() for r in normal)
    june_income = month_total(normal, 6, "INCOME")
    june_expense = month_total(normal, 6, "EXPENSE")
    july_income = month_total(normal, 7, "INCOME")
    july_expense = month_total(normal, 7, "EXPENSE")
    NORMAL_MD.write_text(
        "\n".join([
            "# Excel June/July Normal Truth",
            "",
            f"Excel SHA-256: `{sha}`",
            "",
            "## Sheets inspected",
            *[f"- {name}: {rng}, non-empty rows {first}-{last}" for name, rng, first, last in sheet_info],
            "",
            "## Detected sections",
            "- Normal transactions: T6/T7 row 55 income headers, row 56 variable expense headers, rows 57+ daily detail.",
            "- Fixed expenses: T6/T7 row 45 date buckets, row 46 labels, row 47 amounts.",
            "- Debt/payment: GHI SỔ NỢ left receivable block A:F, right payable block H:M.",
            "- Summary/statistics: TỔNG KẾT NĂM rows 4-60 and T6/T7 rows 1-6, 49-52.",
            "",
            "## Totals",
            f"- June income: {june_income}",
            f"- June expense: {june_expense}",
            f"- June net: {june_income - june_expense}",
            f"- July income: {july_income}",
            f"- July expense: {july_expense}",
            f"- July net: {july_income - july_expense}",
            f"- Combined income: {sum_amount(normal, 'INCOME')}",
            f"- Combined expense: {sum_amount(normal, 'EXPENSE')}",
            f"- Combined net: {sum_amount(normal, 'INCOME') - sum_amount(normal, 'EXPENSE')}",
            "",
            "## Count by category",
            *[f"- {k}: {v}" for k, v in sorted(by_cat.items(), key=lambda kv: normalize_text(kv[0]))],
            "",
            "## Count by day",
            *[f"- {k}: {v}" for k, v in sorted(by_day.items())],
        ]) + "\n",
        encoding="utf-8",
    )
    people = defaultdict(lambda: {"debt_count": 0, "payment_count": 0, "principal": Decimal("0.00"), "paid": Decimal("0.00"), "remaining": Decimal("0.00"), "opened": [], "payments": [], "sources": []})
    for d in debts:
        p = people[d.person_name]
        p["debt_count"] += 1
        p["principal"] += d.principal_amount
        p["paid"] += d.paid_amount
        p["remaining"] += d.remaining_amount
        p["opened"].append(d.opened_date)
        p["sources"].append(f"{d.sheet}!{d.row}")
    for pay in payments:
        p = people[pay.person_name]
        p["payment_count"] += 1
        p["payments"].append(pay.payment_date)
        p["sources"].append(f"{pay.sheet}!{pay.row}")
    found_names = {normalize_text(k): k for k in people}
    PERSON_MD.write_text(
        "\n".join([
            "# Excel June/July Debt Person Summary",
            "",
            f"- Is Dương present? {'yes' if 'duong' in found_names else 'no'}",
            f"- Is mẹ anh present? {'yes' if 'me anh' in found_names else 'no'}",
            f"- Is mẹ em present? {'yes' if 'me em' in found_names else 'no'}",
            f"- Is Thư present? {'yes' if 'thu' in found_names else 'no'}",
            f"- Other people found: {', '.join(sorted(people, key=normalize_text))}",
            "",
            "| personName | debt row count | payment row count | totalPrincipal | totalPaid | totalRemaining | earliestOpenedDate | latestOpenedDate | latestPaymentDate | source sheets/rows |",
            "|---|---:|---:|---:|---:|---:|---|---|---|---|",
            *[
                f"| {name} | {v['debt_count']} | {v['payment_count']} | {v['principal']} | {v['paid']} | {v['remaining']} | {min(v['opened']).isoformat() if v['opened'] else ''} | {max(v['opened']).isoformat() if v['opened'] else ''} | {max(v['payments']).isoformat() if v['payments'] else ''} | {', '.join(v['sources'])} |"
                for name, v in sorted(people.items(), key=lambda kv: normalize_text(kv[0]))
            ],
        ]) + "\n",
        encoding="utf-8",
    )
    DB_NORMAL_MD.write_text("\n".join(["# Excel/DB Normal Diff", "", *[f"- {n}" for n in db["notes"]], f"- DB comparison done: {'yes' if db['done'] else 'no'}", f"- Exact duplicates: {len(db['exact_normal'])}", f"- Likely duplicates: {len(db['likely_normal'])}", f"- Missing normal transactions: {len(db['missing_normal'])}", "- Ambiguous candidates: 0"]) + "\n", encoding="utf-8")
    DB_DEBT_MD.write_text("\n".join(["# Excel/DB Debt Diff", "", *[f"- {n}" for n in db["notes"]], f"- DB comparison done: {'yes' if db['done'] else 'no'}", f"- DB people present: {', '.join(db['db_people']) if db['db_people'] else '(skipped/none)'}", f"- Excel people present: {', '.join(db['excel_people']) if db['excel_people'] else '(none)'}", f"- DB people missing from Excel: {', '.join(db['db_people_missing']) if db['db_people_missing'] else '(none/skipped)'}", f"- Excel people missing from DB: {', '.join(db['excel_people_missing']) if db['excel_people_missing'] else '(none/skipped)'}", f"- Suspicious DB rows with note \"Debt import not part of T1\": {len(db['suspicious'])}", f"- Excel debt rows to compare/import: {len(debts)}", f"- Excel payment rows to compare/import: {len(payments)}"]) + "\n", encoding="utf-8")
    PLAN_MD.write_text(
        "\n".join([
            "# June/July Import Plan",
            "",
            "Mode: dry-run only. No SQL generated. No DB mutation executed.",
            "",
            "## A. Normal transaction import plan",
            f"- rows to import: {len(db['missing_normal'])}",
            f"- rows already in DB: {len(db['exact_normal'])}",
            "- ambiguous rows excluded: 0",
            f"- totals before/after: {'available in DB diff' if db['done'] else 'DB comparison skipped'}",
            "",
            "## B. Debt cleanup/import plan",
            "- existing DB debt rows recommended to keep: requires DB comparison." if not db["done"] else "- existing DB debt rows recommended to keep: see DB debt diff.",
            f"- existing DB debt rows suspicious/old partial: {len(db['suspicious'])}",
            f"- Excel debt rows to import: {len(debts)}",
            f"- Excel payment rows to import: {len(payments)}",
            f"- ambiguous rows excluded: {sum(1 for d in debts if d.ambiguity_reason)} debt rows flagged, not excluded from truth CSV.",
            "",
            "## C. Confirmation required before execution",
            f"- workspace name: {'Tài chính cá nhân của Codex Full History Test'}",
            f"- normal candidate count: {len(db['missing_normal'])}",
            f"- normal income total: {sum_amount(db['missing_normal'], 'INCOME')}",
            f"- normal expense total: {sum_amount(db['missing_normal'], 'EXPENSE')}",
            f"- debt person count: {len({d.person_name for d in debts})}",
            f"- debt principal total: {sum((d.principal_amount for d in debts), Decimal('0.00'))}",
            f"- debt payment total: {sum((p.amount for p in payments), Decimal('0.00'))}",
            "- whether old partial debt rows should be deleted/replaced: requires explicit confirmation",
        ]) + "\n",
        encoding="utf-8",
    )


def write_reports(excel: Path, workspace_name: str) -> dict:
    REPORT_DIR.mkdir(exist_ok=True)
    wb, sheet_info = inspect_workbook(excel)
    normal = parse_normal_transactions(wb)
    debts, payments = parse_debt_truth(wb)
    db = db_compare(workspace_name, normal, debts, payments)
    write_normal_csv(normal)
    write_debt_csv(debts)
    write_payment_csv(payments)
    write_markdown(excel, sheet_info, normal, debts, payments, db)
    return {"normal": normal, "debts": debts, "payments": payments, "db": db, "sheets": sheet_info}


def build_parser() -> argparse.ArgumentParser:
    p = argparse.ArgumentParser()
    p.add_argument("--excel", required=True, type=Path)
    p.add_argument("--workspace-name", required=True)
    p.add_argument("--dry-run", action="store_true", default=True)
    p.add_argument("--prepare-sql", action="store_true")
    p.add_argument("--execute", action="store_true")
    p.add_argument("--i-understand-this-mutates-db", action="store_true")
    p.add_argument("--confirm-workspace-name")
    p.add_argument("--confirm-normal-count", type=int)
    p.add_argument("--confirm-normal-expense-total", type=Decimal)
    p.add_argument("--confirm-normal-income-total", type=Decimal)
    return p


def main(argv: list[str] | None = None) -> int:
    args = build_parser().parse_args(argv)
    excel = workbook_path(args.excel)
    result = write_reports(excel, args.workspace_name)
    ensure_execute_guard(args, result["db"]["missing_normal"])
    if args.execute:
        raise SystemExit("execute exists but is intentionally not implemented for this task")
    if args.prepare_sql:
        raise SystemExit("prepare-sql is intentionally disabled until mappings are reviewed")
    print(f"normal={len(result['normal'])} debt={len(result['debts'])} payments={len(result['payments'])} db_compare={'yes' if result['db']['done'] else 'no'}")
    print(f"reports={REPORT_DIR}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
